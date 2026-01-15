#!/usr/bin/env python3
"""
Automated Excel Validation Tool for Comparison Reports
Usage: python3 validate_excel.py <excel_file> <section_name>
"""

import openpyxl
import sys
import os

def validate_excel(excel_file: str, section_name: str) -> bool:
    """
    Automated Excel validation
    Returns: True if all checks pass, False otherwise
    """
    checks_passed = 0
    checks_total = 0
    
    try:
        print(f"[VALIDATE] Opening {excel_file}...")
        wb = openpyxl.load_workbook(excel_file)
        checks_total += 1
        checks_passed += 1
        print("[OK] Excel file opens without corruption")
        
        # Check required tabs exist
        checks_total += 1
        if "Summary Dashboard" in wb.sheetnames:
            checks_passed += 1
            print("[OK] Summary Dashboard tab exists")
        else:
            print("[ERROR] Summary Dashboard tab missing")
        
        checks_total += 1
        expected_tab = f"{section_name.title()} Comparison"
        if expected_tab in wb.sheetnames or any(section_name.lower() in name.lower() for name in wb.sheetnames):
            checks_passed += 1
            print(f"[OK] Section tab exists")
        else:
            print(f"[ERROR] Section tab missing: {expected_tab}")
            return False
        
        # Find the section tab
        ws = None
        for name in wb.sheetnames:
            if section_name.lower() in name.lower() and name != "Summary Dashboard":
                ws = wb[name]
                break
        
        if ws is None:
            print(f"[ERROR] Could not find section worksheet")
            return False
        
        # Check tab has data
        checks_total += 1
        if ws.max_row > 1:
            checks_passed += 1
            print(f"[OK] Tab has data ({ws.max_row} rows)")
        else:
            print("[ERROR] Tab is empty")
        
        # Check headers exist
        headers = [cell.value for cell in ws[1] if cell.value]
        checks_total += 1
        required_headers = ["Metric", "Reference", "Test"]
        if all(any(req.lower() in str(h).lower() for h in headers) for req in required_headers):
            checks_passed += 1
            print(f"[OK] Required headers present: {headers}")
        else:
            print(f"[ERROR] Missing required headers. Found: {headers}")
        
        # Check for formula errors
        checks_total += 1
        formula_errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell.value, str) and cell.value.startswith('#'):
                    formula_errors.append(f"Row {row_idx}, Col {col_idx}: {cell.value}")
        
        if not formula_errors:
            checks_passed += 1
            print("[OK] No formula errors found")
        else:
            print(f"[ERROR] Formula errors found: {formula_errors[:3]}")
        
        # Summary
        print(f"\n[SUMMARY] Passed {checks_passed}/{checks_total} automated checks")
        
        if checks_passed == checks_total:
            print("[OK] All automated validation checks PASSED")
            print("[INFO] Manual verification still required (formatting, colors, UX)")
            return True
        else:
            print(f"[ERROR] {checks_total - checks_passed} checks FAILED")
            return False
            
    except Exception as e:
        print(f"[ERROR] Excel validation exception: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 validate_excel.py <excel_file> <section_name>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    section_name = sys.argv[2]
    
    if not os.path.exists(excel_file):
        print(f"[ERROR] File not found: {excel_file}")
        sys.exit(1)
    
    success = validate_excel(excel_file, section_name)
    sys.exit(0 if success else 1)



